from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from pprint import pprint as pp

wb = Workbook()
dest_filename = 'test_wb2.xlsx'

ws1 = wb.active
ws1.title = "range names"

## add rows 
for row in range(1, 40):
    ws1.append(range(600))

## create new sheet
ws2 = wb.create_sheet(title="Pi")

## specify a cell value
ws2['F5'] = 3.14

## create new sheet and add some data in a given range
## row and columns below show how they can be referred to
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
     for col in range(27, 54):
       cells = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
ws3.column_dimensions.group('F','AD', hidden=True)
wb.save(filename = dest_filename)

## Prbably dont
